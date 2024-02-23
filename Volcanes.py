import pandas as pd
from openpyxl import load_workbook
import datetime
import os

import locale

# Establecer el locale a español
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# Carga de datos
model_galeras = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 10, nrows = 1, sep = '\s+')
model_huila = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 17, nrows = 1, sep = '\s+')
model_machin = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 24, nrows = 1, sep = '\s+')
model_ruiz = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 31, nrows = 1, sep = '\s+')
model_purace = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 38, nrows = 1, sep = '\s+')
model_azufral = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 45, nrows = 1, sep = '\s+')
model_chiles = pd.read_csv("http://bart.ideam.gov.co/wrfideam/new_modelo/WRF00COLOMBIA/txt/VOLCAN.txt", skiprows = 52, nrows = 1, sep = '\s+')

# Procesamiento de datos
volcanes_u = model_galeras.append(model_huila).append(model_machin).append(model_ruiz).append(model_chiles)
volcanes_u["Volcanes"] = ["VOLCAN GALERAS", "VOLCAN NEVADO DEL HUILA", "VOLCAN CERRO MACHIN", "VOLCAN NEVADO DEL RUIZ", "AREA DEL COMPLEJO CERRO NEGRO DE MAYASQUER"]

volcanes_u = pd.pivot_table(volcanes_u, index = "Volcanes", values = ["D10M", "V10M", "D18M", "V18M", "D24M", "V24M", "D30M", "V30M"] )
volcanes_u = volcanes_u[["D10M", "V10M", "D18M", "V18M", "D24M", "V24M", "D30M", "V30M"]]

# Definición de funciones
def df(dataframe):
    if dataframe >= 0 and dataframe <= 22:
        val = "NORTE"
    elif dataframe >= 23 and dataframe <= 67:
        val = "NORESTE"
    elif dataframe >= 68 and dataframe <= 112:
        val = "ESTE"
    elif dataframe >= 113 and dataframe <= 157:
        val = "SURESTE"
    elif dataframe >= 158 and dataframe <= 202:
        val = "SUR"
    elif dataframe >= 203 and dataframe <= 247:
        val = "SUROESTE"
    elif dataframe >= 248 and dataframe <= 292:
        val = "OESTE"
    elif dataframe >= 293 and dataframe <= 360:
        val = "NOROESTE"
    return val

def velocidades(dataframe):
    if dataframe >= 0 and dataframe < 5:
        val = "0-5"
    elif dataframe >= 5 and dataframe < 10:
        val = "5-10"
    elif dataframe >= 10 and dataframe < 15:
        val = "10-15"
    elif dataframe >= 15 and dataframe < 20:
        val = "15-20"
    elif dataframe >= 20 and dataframe < 25:
        val = "20-25"
    elif dataframe >= 25 and dataframe < 30:
        val = "25-30"
    elif dataframe >= 30 and dataframe < 35:
        val = "30-35"
    elif dataframe >= 35 and dataframe < 40:
        val = "35-40"
    elif dataframe >= 40 and dataframe < 50:
        val = "40-50"
    elif dataframe >= 50 and dataframe < 60:
        val = "50-60"
    return val

def velocidades_km(dataframe):
    if dataframe >= 0 and dataframe < 5:
        val = "0-9"
    elif dataframe >= 5 and dataframe < 10:
        val = "9-18"
    elif dataframe >= 10 and dataframe < 15:
        val = "18-27"
    elif dataframe >= 15 and dataframe < 20:
        val = "27-36"
    elif dataframe >= 20 and dataframe < 25:
        val = "36-45"
    elif dataframe >= 25 and dataframe < 30:
        val = "45-54"
    elif dataframe >= 30 and dataframe < 35:
        val = "54-63"
    elif dataframe >= 35 and dataframe < 40:
        val = "63-72"
    elif dataframe >= 40 and dataframe < 50:
        val = "72-90"
    elif dataframe >= 50 and dataframe < 60:
        val = "90 - 108"
    return val

# Aplicación de las funciones
volcanes_u["D10M_C"] = volcanes_u["D10M"].apply(df)
volcanes_u["D18M_C"] = volcanes_u["D18M"].apply(df)
volcanes_u["D24M_C"] = volcanes_u["D24M"].apply(df)
volcanes_u["D30M_C"] = volcanes_u["D30M"].apply(df)

volcanes_u["V10M_C"] = volcanes_u["V10M"].apply(velocidades)
volcanes_u["V18M_C"] = volcanes_u["V18M"].apply(velocidades)
volcanes_u["V24M_C"] = volcanes_u["V24M"].apply(velocidades)
volcanes_u["V30M_C"] = volcanes_u["V30M"].apply(velocidades)

volcanes_u["V10M_KM"] = volcanes_u["V10M"].apply(velocidades_km)
volcanes_u["V18M_KM"] = volcanes_u["V18M"].apply(velocidades_km)
volcanes_u["V24M_KM"] = volcanes_u["V24M"].apply(velocidades_km)
volcanes_u["V30M_KM"] = volcanes_u["V30M"].apply(velocidades_km)

volcanes_f = volcanes_u[['D10M_C', 'V10M_C', 'V10M_KM', 'D18M_C', 'V18M_C', 'V18M_KM', 'D24M_C', 'V24M_C', 'V24M_KM', 'D30M_C', 'V30M_C', 'V30M_KM']]

# Guardar archivo
fecha_actual = datetime.datetime.now()
nombre_mes = fecha_actual.strftime("%B_%Y")  # Enero_2024
nombre_archivo = fecha_actual.strftime("%Y-%m-%d") + "_volcanes.xlsx"  # 2024-01-01_volcanes.xlsx

ruta_carpeta = os.path.join(r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\Volcanes", nombre_mes)
os.makedirs(ruta_carpeta, exist_ok=True)

ruta_completa_archivo = os.path.join(ruta_carpeta, nombre_archivo)
volcanes_f.to_excel(ruta_completa_archivo)

# Continuación del código original para manejar workbooks de Excel...

# In[39]:



# Cargar workbook existente
fecha_actual = datetime.datetime.now()
nombre_mes = fecha_actual.strftime("%B_%Y")  # 'enero_2024' por ejemplo
ruta_carpeta = os.path.join(r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\Volcanes", nombre_mes)
ruta_salida = os.path.join(r"O:\Mi unidad\OSPA\01. Tematicas\01. Meteorologia\01. Productos\10. Volcanes", nombre_mes)
os.makedirs(ruta_carpeta, exist_ok=True)  # Crear la carpeta si no existe
os.makedirs(ruta_salida, exist_ok=True)  # Crear la carpeta si no existe


nombre_archivo_volcanes = fecha_actual.strftime("%Y-%m-%d") + "_volcanes.xlsx"
ruta_archivo_volcanes = os.path.join(ruta_carpeta, nombre_archivo_volcanes)

volcanes = load_workbook(filename = ruta_archivo_volcanes)
# Asegúrate de actualizar la ruta y nombre del archivo de pronóstico según sea necesario
pronostico = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\Volcanes\2022-01-28_PronVolcanes.xlsx")

h_pronostico = pronostico["Hoja1"]
h_volcanes = volcanes["Sheet1"]

# Función para llenar datos en el pronóstico
def llenar_volcanes(p, v):
   #D10M
    p.cell(row = 10, column = 3).value = v.cell(row = 6, column = 2).value
    #D18M
    p.cell(row = 11, column = 3).value = v.cell(row = 6, column = 5).value
    #D24M
    p.cell(row = 12, column = 3).value = v.cell(row = 6, column = 8).value
    #D30M
    p.cell(row = 13, column = 3).value = v.cell(row = 6, column = 11).value
    
    #V10M
    p.cell(row = 10, column = 4).value = v.cell(row = 6, column = 3).value
    #V18M
    p.cell(row = 11, column = 4).value = v.cell(row = 6, column = 6).value
    #V24M
    p.cell(row = 12, column = 4).value = v.cell(row = 6, column = 9).value
    #V30M
    p.cell(row = 13, column = 4).value = v.cell(row = 6, column = 12).value
    
    #V10KM
    p.cell(row = 10, column = 5).value = v.cell(row = 6, column = 4).value
    #V18KM
    p.cell(row = 11, column = 5).value = v.cell(row = 6, column = 7).value
    #V24KM
    p.cell(row = 12, column = 5).value = v.cell(row = 6, column = 10).value
    #V30KM
    p.cell(row = 13, column = 5).value = v.cell(row = 6, column = 13).value
    
    ##NEVADO DEL HUILA
    
    #D10M
    p.cell(row = 22, column = 3).value = v.cell(row = 5, column = 2).value
    #D18M
    p.cell(row = 23, column = 3).value = v.cell(row = 5, column = 5).value
    #D24M
    p.cell(row = 24, column = 3).value = v.cell(row = 5, column = 8).value
    #D30M
    p.cell(row = 25, column = 3).value = v.cell(row = 5, column = 11).value
    
    #V10M
    p.cell(row = 22, column = 4).value = v.cell(row = 5, column = 3).value
    #V18M
    p.cell(row = 23, column = 4).value = v.cell(row = 5, column = 6).value
    #V24M
    p.cell(row = 24, column = 4).value = v.cell(row = 5, column = 9).value
    #V30M
    p.cell(row = 25, column = 4).value = v.cell(row = 5, column = 12).value
    
    #V10KM
    p.cell(row = 22, column = 5).value = v.cell(row = 5, column = 4).value
    #V18KM
    p.cell(row = 23, column = 5).value = v.cell(row = 5, column = 7).value
    #V24KM
    p.cell(row = 24, column = 5).value = v.cell(row = 5, column = 10).value
    #V30KM
    p.cell(row = 25, column = 5).value = v.cell(row = 5, column = 13).value
    
    ##GALERAS
    
    #D10M
    p.cell(row = 46, column = 3).value = v.cell(row = 4, column = 2).value
    #D18M
    p.cell(row = 47, column = 3).value = v.cell(row = 4, column = 5).value
    #D24M
    p.cell(row = 48, column = 3).value = v.cell(row = 4, column = 8).value
    #D30M
    p.cell(row = 49, column = 3).value = v.cell(row = 4, column = 11).value
    
    #V10M
    p.cell(row = 46, column = 4).value = v.cell(row = 4, column = 3).value
    #V18M
    p.cell(row = 47, column = 4).value = v.cell(row = 4, column = 6).value
    #V24M
    p.cell(row = 48, column = 4).value = v.cell(row = 4, column = 9).value
    #V30M
    p.cell(row = 49, column = 4).value = v.cell(row = 4, column = 12).value
    
    #V10KM
    p.cell(row = 46, column = 5).value = v.cell(row = 4, column = 4).value
    #V18KM
    p.cell(row = 47, column = 5).value = v.cell(row = 4, column = 7).value
    #V24KM
    p.cell(row = 48, column = 5).value = v.cell(row = 4, column = 10).value
    #V30KM
    p.cell(row = 49, column = 5).value = v.cell(row = 4, column = 13).value
    
    ##MACHIN
    
    #D10M
    p.cell(row = 34, column = 3).value = v.cell(row = 3, column = 2).value
    #D18M
    p.cell(row = 35, column = 3).value = v.cell(row = 3, column = 5).value
    #D24M
    p.cell(row = 36, column = 3).value = v.cell(row = 3, column = 8).value
    #D30M
    p.cell(row = 37, column = 3).value = v.cell(row = 3, column = 11).value
    
    #V10M
    p.cell(row = 34, column = 4).value = v.cell(row = 3, column = 3).value
    #V18M
    p.cell(row = 35, column = 4).value = v.cell(row = 3, column = 6).value
    #V24M
    p.cell(row = 36, column = 4).value = v.cell(row = 3, column = 9).value
    #V30M
    p.cell(row = 37, column = 4).value = v.cell(row = 3, column = 12).value
    
    #V10KM
    p.cell(row = 34, column = 5).value = v.cell(row = 3, column = 4).value
    #V18KM
    p.cell(row = 35, column = 5).value = v.cell(row = 3, column = 7).value
    #V24KM
    p.cell(row = 36, column = 5).value = v.cell(row = 3, column = 10).value
    #V30KM
    p.cell(row = 37, column = 5).value = v.cell(row = 3, column = 13).value
    
    ##MAYASQUER
    
    #D10M
    p.cell(row = 58, column = 3).value = v.cell(row = 2, column = 2).value
    #D18M
    p.cell(row = 59, column = 3).value = v.cell(row = 2, column = 5).value
    #D24M
    p.cell(row = 60, column = 3).value = v.cell(row = 2, column = 8).value
    #D30M
    p.cell(row = 61, column = 3).value = v.cell(row = 2, column = 11).value
    
    #V10M
    p.cell(row = 58, column = 4).value = v.cell(row = 2, column = 3).value
    #V18M
    p.cell(row = 59, column = 4).value = v.cell(row = 2, column = 6).value
    #V24M
    p.cell(row = 60, column = 4).value = v.cell(row = 2, column = 9).value
    #V30M
    p.cell(row = 61, column = 4).value = v.cell(row = 2, column = 12).value
    
    #V10KM
    p.cell(row = 58, column = 5).value = v.cell(row = 2, column = 4).value
    #V18KM
    p.cell(row = 59, column = 5).value = v.cell(row = 2, column = 7).value
    #V24KM
    p.cell(row = 60, column = 5).value = v.cell(row = 2, column = 10).value
    #V30KM
    p.cell(row = 61, column = 5).value = v.cell(row = 2, column = 13).value
    
    # Guardar el archivo de pronóstico en la carpeta correspondiente
    nombre_archivo_pronostico = fecha_actual.strftime("%Y-%m-%d") + "_PronVolcanes.xlsx"
    ruta_archivo_pronostico = os.path.join(ruta_salida, nombre_archivo_pronostico)
    return pronostico.save(ruta_archivo_pronostico)

# Llamada a la función para realizar la actualización
correr_volcanes = llenar_volcanes(h_pronostico, h_volcanes)




# volcanes = load_workbook(filename = fr"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\Volcanes\{datetime.datetime.now().strftime('%Y-%m-%d')}_volcanes.xlsx")
# pronostico = load_workbook(filename = r"C:\Users\easalazarm\Documents\IDEAM\IDEAM\Codigos\Volcanes\2022-01-28_PronVolcanes.xlsx")

# h_pronostico = pronostico["Hoja1"]
# h_volcanes = volcanes["Sheet1"]


# # In[44]:


# def llenar_volcanes(p, v):
#     ##NEVADO DEL RUIZ 
    
#     #D10M
#     p.cell(row = 10, column = 3).value = v.cell(row = 6, column = 2).value
#     #D18M
#     p.cell(row = 11, column = 3).value = v.cell(row = 6, column = 5).value
#     #D24M
#     p.cell(row = 12, column = 3).value = v.cell(row = 6, column = 8).value
#     #D30M
#     p.cell(row = 13, column = 3).value = v.cell(row = 6, column = 11).value
    
#     #V10M
#     p.cell(row = 10, column = 4).value = v.cell(row = 6, column = 3).value
#     #V18M
#     p.cell(row = 11, column = 4).value = v.cell(row = 6, column = 6).value
#     #V24M
#     p.cell(row = 12, column = 4).value = v.cell(row = 6, column = 9).value
#     #V30M
#     p.cell(row = 13, column = 4).value = v.cell(row = 6, column = 12).value
    
#     #V10KM
#     p.cell(row = 10, column = 5).value = v.cell(row = 6, column = 4).value
#     #V18KM
#     p.cell(row = 11, column = 5).value = v.cell(row = 6, column = 7).value
#     #V24KM
#     p.cell(row = 12, column = 5).value = v.cell(row = 6, column = 10).value
#     #V30KM
#     p.cell(row = 13, column = 5).value = v.cell(row = 6, column = 13).value
    
#     ##NEVADO DEL HUILA
    
#     #D10M
#     p.cell(row = 22, column = 3).value = v.cell(row = 5, column = 2).value
#     #D18M
#     p.cell(row = 23, column = 3).value = v.cell(row = 5, column = 5).value
#     #D24M
#     p.cell(row = 24, column = 3).value = v.cell(row = 5, column = 8).value
#     #D30M
#     p.cell(row = 25, column = 3).value = v.cell(row = 5, column = 11).value
    
#     #V10M
#     p.cell(row = 22, column = 4).value = v.cell(row = 5, column = 3).value
#     #V18M
#     p.cell(row = 23, column = 4).value = v.cell(row = 5, column = 6).value
#     #V24M
#     p.cell(row = 24, column = 4).value = v.cell(row = 5, column = 9).value
#     #V30M
#     p.cell(row = 25, column = 4).value = v.cell(row = 5, column = 12).value
    
#     #V10KM
#     p.cell(row = 22, column = 5).value = v.cell(row = 5, column = 4).value
#     #V18KM
#     p.cell(row = 23, column = 5).value = v.cell(row = 5, column = 7).value
#     #V24KM
#     p.cell(row = 24, column = 5).value = v.cell(row = 5, column = 10).value
#     #V30KM
#     p.cell(row = 25, column = 5).value = v.cell(row = 5, column = 13).value
    
#     ##GALERAS
    
#     #D10M
#     p.cell(row = 46, column = 3).value = v.cell(row = 4, column = 2).value
#     #D18M
#     p.cell(row = 47, column = 3).value = v.cell(row = 4, column = 5).value
#     #D24M
#     p.cell(row = 48, column = 3).value = v.cell(row = 4, column = 8).value
#     #D30M
#     p.cell(row = 49, column = 3).value = v.cell(row = 4, column = 11).value
    
#     #V10M
#     p.cell(row = 46, column = 4).value = v.cell(row = 4, column = 3).value
#     #V18M
#     p.cell(row = 47, column = 4).value = v.cell(row = 4, column = 6).value
#     #V24M
#     p.cell(row = 48, column = 4).value = v.cell(row = 4, column = 9).value
#     #V30M
#     p.cell(row = 49, column = 4).value = v.cell(row = 4, column = 12).value
    
#     #V10KM
#     p.cell(row = 46, column = 5).value = v.cell(row = 4, column = 4).value
#     #V18KM
#     p.cell(row = 47, column = 5).value = v.cell(row = 4, column = 7).value
#     #V24KM
#     p.cell(row = 48, column = 5).value = v.cell(row = 4, column = 10).value
#     #V30KM
#     p.cell(row = 49, column = 5).value = v.cell(row = 4, column = 13).value
    
#     ##MACHIN
    
#     #D10M
#     p.cell(row = 34, column = 3).value = v.cell(row = 3, column = 2).value
#     #D18M
#     p.cell(row = 35, column = 3).value = v.cell(row = 3, column = 5).value
#     #D24M
#     p.cell(row = 36, column = 3).value = v.cell(row = 3, column = 8).value
#     #D30M
#     p.cell(row = 37, column = 3).value = v.cell(row = 3, column = 11).value
    
#     #V10M
#     p.cell(row = 34, column = 4).value = v.cell(row = 3, column = 3).value
#     #V18M
#     p.cell(row = 35, column = 4).value = v.cell(row = 3, column = 6).value
#     #V24M
#     p.cell(row = 36, column = 4).value = v.cell(row = 3, column = 9).value
#     #V30M
#     p.cell(row = 37, column = 4).value = v.cell(row = 3, column = 12).value
    
#     #V10KM
#     p.cell(row = 34, column = 5).value = v.cell(row = 3, column = 4).value
#     #V18KM
#     p.cell(row = 35, column = 5).value = v.cell(row = 3, column = 7).value
#     #V24KM
#     p.cell(row = 36, column = 5).value = v.cell(row = 3, column = 10).value
#     #V30KM
#     p.cell(row = 37, column = 5).value = v.cell(row = 3, column = 13).value
    
#     ##MAYASQUER
    
#     #D10M
#     p.cell(row = 58, column = 3).value = v.cell(row = 2, column = 2).value
#     #D18M
#     p.cell(row = 59, column = 3).value = v.cell(row = 2, column = 5).value
#     #D24M
#     p.cell(row = 60, column = 3).value = v.cell(row = 2, column = 8).value
#     #D30M
#     p.cell(row = 61, column = 3).value = v.cell(row = 2, column = 11).value
    
#     #V10M
#     p.cell(row = 58, column = 4).value = v.cell(row = 2, column = 3).value
#     #V18M
#     p.cell(row = 59, column = 4).value = v.cell(row = 2, column = 6).value
#     #V24M
#     p.cell(row = 60, column = 4).value = v.cell(row = 2, column = 9).value
#     #V30M
#     p.cell(row = 61, column = 4).value = v.cell(row = 2, column = 12).value
    
#     #V10KM
#     p.cell(row = 58, column = 5).value = v.cell(row = 2, column = 4).value
#     #V18KM
#     p.cell(row = 59, column = 5).value = v.cell(row = 2, column = 7).value
#     #V24KM
#     p.cell(row = 60, column = 5).value = v.cell(row = 2, column = 10).value
#     #V30KM
#     p.cell(row = 61, column = 5).value = v.cell(row = 2, column = 13).value
    
#     # return pronostico.save(filename = fr"M:\OF_SERVICIO_DE_PRONOSTICO_Y_ALERTAS\Compartida\2.Análisis_pronóstico_del_tiempo\2.5_Volcanes\{datetime.datetime.now().strftime('%Y-%m-%d')}_PronVolcanes.xlsx")
#     return pronostico.save(filename = fr"O:\Mi unidad\OSPA\01. Tematicas\01. Meteorologia\01. Productos\10. Volcanes\10.Octubre\{datetime.datetime.now().strftime('%Y-%m-%d')}_PronVolcanes.xlsx")

# correr_volcanes = llenar_volcanes(h_pronostico, h_volcanes)
# correr_volcanes


# # In[40]:





# # In[41]:





# # In[ ]:




